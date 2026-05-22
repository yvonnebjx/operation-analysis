import openpyxl
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.line import LineProperties

def analyze_mgmt_report(input_file, output_file, month_cutoff=11):
    print(f"Loading {input_file}...")
    wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    
    # Target sheet
    sheet_name = '实际数据'
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    
    # Find headers
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {h: i for i, h in enumerate(headers) if h}
    
    # Map required fields
    IDX_DATE = header_map.get('日期')
    IDX_CAT = header_map.get('三级分类') or header_map.get('三级品类')
    IDX_GMV = header_map.get('GMV')
    IDX_PROFIT = header_map.get('销售毛利')
    
    if None in [IDX_DATE, IDX_CAT, IDX_GMV, IDX_PROFIT]:
        print(f"Error: Missing columns. Found: {list(header_map.keys())}")
        return

    data_24 = {}
    data_23 = {}
    
    print("Processing rows...")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        date = row[IDX_DATE]
        cat = row[IDX_CAT]
        gmv = row[IDX_GMV] or 0.0
        profit = row[IDX_PROFIT] or 0.0
        
        if not isinstance(date, datetime):
            continue
            
        if date.year == 2024 and date.month <= month_cutoff:
            target = data_24
        elif date.year == 2023 and date.month <= month_cutoff:
            target = data_23
        else:
            continue
            
        if cat not in target:
            target[cat] = {'gmv': 0.0, 'profit': 0.0}
        target[cat]['gmv'] += float(gmv)
        target[cat]['profit'] += float(profit)

    # Compile results
    all_cats = set(data_24.keys()) | set(data_23.keys())
    results = []
    
    # Grouping for summary
    groups = {
        "盈利&GMV下滑": [], 
        "盈利&GMV增长": [], 
        "亏损&GMV增长": [], 
        "亏损&GMV下滑": []
    }
    
    # Data for scatter plot
    scatter_data = []
    
    for cat in all_cats:
        d24 = data_24.get(cat, {'gmv': 0.0, 'profit': 0.0})
        d23 = data_23.get(cat, {'gmv': 0.0, 'profit': 0.0})
        
        gmv_24 = d24['gmv']
        gmv_23 = d23['gmv']
        profit_24 = d24['profit']
        profit_23 = d23['profit']
        
        gmv_yoy = (gmv_24 - gmv_23) / gmv_23 if gmv_23 != 0 else 0.0
        gp_24 = profit_24 / gmv_24 if gmv_24 != 0 else 0.0
        gp_23 = profit_23 / gmv_23 if gmv_23 != 0 else 0.0
        
        s1 = '+' if gp_24 > gp_23 else '-'
        s2 = '+' if profit_24 > profit_23 else '-'
        label = f"{cat}{s1}{s2}"
        
        profit_status = "盈利" if gp_24 >= 0.08 else "亏损"
        growth_status = "增长" if gmv_yoy >= 0 else "下滑"
        quad_name = f"{profit_status}&GMV{growth_status}"
            
        results.append([
            cat, gmv_24, gmv_23, gmv_yoy, profit_24, profit_23, gp_24, quad_name
        ])
        
        if quad_name in groups:
            groups[quad_name].append(label)
            
        if gmv_24 > 0:
            # Clip values: GP%: [-40%, 40%], GMV YoY: [-100%, 200%]
            plot_gp = max(min(gp_24, 0.4), -0.4)
            plot_yoy = max(min(gmv_yoy, 2.0), -1.0)
            scatter_data.append([cat, plot_gp, plot_yoy, abs(profit_24)])
        
    results.sort(key=lambda x: x[1], reverse=True)
    
    out_wb = Workbook()
    
    # 1. Summary Sheet
    ws_sum = out_wb.active
    ws_sum.title = "结论汇总"
    summary_headers = [
        ("第【1】区", "盈利&GMV下滑", "FFC000"),
        ("第【2】区", "盈利&GMV增长", "92D050"),
        ("第【3】区", "亏损&GMV增长", "FFD966"),
        ("第【4】区", "亏损&GMV下滑", "FF0000")
    ]
    
    for i, (h1, h2, color) in enumerate(summary_headers, 1):
        c1 = ws_sum.cell(row=1, column=i, value=h1)
        c2 = ws_sum.cell(row=2, column=i, value=h2)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font_style = Font(bold=True, color="FFFFFF") if color == "FF0000" else Font(bold=True)
        c1.fill = fill
        c1.font = font_style
        c1.alignment = Alignment(horizontal="center")
        c2.fill = fill
        c2.font = font_style
        c2.alignment = Alignment(horizontal="center")
        
        q_items = groups[h2]
        for r_idx, label in enumerate(q_items, 3):
            cell = ws_sum.cell(row=r_idx, column=i, value=label)
            cell.alignment = Alignment(horizontal="center")
        ws_sum.column_dimensions[get_column_letter(i)].width = 25

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    # 2. Detail Sheet
    ws_det = out_wb.create_sheet("明细数据")
    det_headers = ['三级分类', '2024 YTD GMV', '2023 YTD GMV', 'GMV YoY', '2024 YTD 毛利', '2023 YTD 毛利', '2024 GP%', '四象限分类']
    ws_det.append(det_headers)
    for r in results:
        ws_det.append(r)
    
    for row in ws_det.iter_rows(min_row=2):
        row[3].number_format = '0.00%'
        row[6].number_format = '0.00%'
        row[1].number_format = '#,##0'
        row[4].number_format = '#,##0'
    
    for col in ws_det.columns:
        ws_det.column_dimensions[col[0].column_letter].width = 20

    # 3. Scatter Chart Sheet
    ws_chart_data = out_wb.create_sheet("图表数据")
    chart_headers = ['Category', 'GP%', 'GMV YoY']
    ws_chart_data.append(chart_headers)
    for s in sorted(scatter_data, key=lambda x: data_24[x[0]]['gmv'], reverse=True)[:50]:
        ws_chart_data.append([s[0], s[1], s[2]])
        
    for row in ws_chart_data.iter_rows(min_row=2):
        row[1].number_format = '0.00%'
        row[2].number_format = '0.00%'

    # Create Scatter Chart
    chart = ScatterChart()
    chart.style = 13
    chart.title = f"GMV和GP增长矩阵-2024年{month_cutoff}月YTD"
    
    # Blue Dashed Line Properties
    ln = LineProperties(prstDash='dash', w=19050)
    ln.solidFill = ColorChoice(prstClr='blue')
    
    # 1. Horizontal Axis (X) — mapped to GMV YoY settings (WPS swap workaround)
    chart.x_axis.title = 'GMV YoY增长率'
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.scaling.min = -1.0
    chart.x_axis.scaling.max = 2.0
    chart.x_axis.majorUnit = 0.5
    chart.x_axis.crossesAt = 0
    chart.x_axis.spPr = GraphicalProperties(ln=ln)
    
    # 2. Vertical Axis (Y) — mapped to GP% settings (WPS swap workaround)
    chart.y_axis.title = '24年YTD-GP%'
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.scaling.min = -0.32
    chart.y_axis.scaling.max = 0.40
    chart.y_axis.majorUnit = 0.08
    chart.y_axis.crossesAt = 0.08
    chart.y_axis.spPr = GraphicalProperties(ln=ln)
    
    # Chart Size
    chart.width = 32
    chart.height = 16
    
    # Critical fix for openpyxl 3.1.4+ chart rendering bug (axes disappearing in WPS/Excel)
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    
    # Explicit plot area layout to prevent title/label overlap in WPS
    from openpyxl.chart.layout import Layout, ManualLayout
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.005, y=0.05,
            w=0.75, h=0.8,
            xMode="factor", yMode="factor",
            wMode="factor", hMode="factor"
        )
    )
    chart.layout.layoutTarget = "inner"

    xvalues = Reference(ws_chart_data, min_col=2, min_row=2, max_row=ws_chart_data.max_row)
    yvalues = Reference(ws_chart_data, min_col=3, min_row=2, max_row=ws_chart_data.max_row)
    
    series = Series(values=yvalues, xvalues=xvalues, title="Category")
    series.marker.symbol = "circle"
    series.marker.size = 8
    
    # Marker styling
    marker_spPr = GraphicalProperties(solidFill=ColorChoice(prstClr='darkBlue'))
    marker_spPr.ln = LineProperties(solidFill=ColorChoice(prstClr='white'), w=5000)
    series.marker.spPr = marker_spPr
    
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    
    from openpyxl.drawing.text import RichTextProperties
    chart.x_axis.title.textProperties = RichTextProperties(vert="horz")
    chart.y_axis.title.textProperties = RichTextProperties(vert="horz")
    
    ws_graph = out_wb.create_sheet("经营情况矩阵图")
    ws_graph.add_chart(chart, "A1")
    
    out_wb.save(output_file)
    print(f"Analysis saved to {output_file}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--month', type=int, default=11)
    args = parser.parse_args()
    analyze_mgmt_report(args.input, args.output, args.month)
