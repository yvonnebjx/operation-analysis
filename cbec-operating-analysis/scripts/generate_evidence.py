from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from run_paths import resolve_run_root, stage_dir


def load_mapping_memory(mapping_path: Path) -> dict[str, str]:
    text = mapping_path.read_text(encoding="utf-8")
    mappings: dict[str, str] = {}
    for source, target in re.findall(r"`([^`]+)` -> `([^`]+)`", text):
        mappings[source] = target
    return mappings


def normalize_header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def normalize_text(value: object, default: str = "未知") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def classify_headers(headers: list[str], mappings: dict[str, str]) -> tuple[list[tuple[str, str]], list[str]]:
    confirmed: list[tuple[str, str]] = []
    unknown: list[str] = []
    for header in headers:
        if not header:
            continue
        if header in mappings:
            confirmed.append((header, mappings[header]))
        else:
            unknown.append(header)
    return confirmed, unknown


def workbook_profile(workbook_path: Path, mappings: dict[str, str]) -> tuple[list[str], dict[str, list[str]], dict[str, list[tuple[str, str]]], dict[str, list[str]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    headers_by_sheet: dict[str, list[str]] = {}
    confirmed_by_sheet: dict[str, list[tuple[str, str]]] = {}
    unknown_by_sheet: dict[str, list[str]] = {}
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [normalize_header(cell) for cell in first_row if normalize_header(cell)]
        headers_by_sheet[sheet_name] = headers
        confirmed, unknown = classify_headers(headers, mappings)
        confirmed_by_sheet[sheet_name] = confirmed
        unknown_by_sheet[sheet_name] = unknown
    return sheet_names, headers_by_sheet, confirmed_by_sheet, unknown_by_sheet


def sheet_date_range(workbook_path: Path) -> dict[str, str | None]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    result: dict[str, str | None] = {
        "actual_min": None,
        "actual_max": None,
        "budget_min": None,
        "budget_max": None,
    }

    def scan_sheet(sheet_name: str, header_name: str, min_key: str, max_key: str) -> None:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, ())
        idx = {name: i for i, name in enumerate(headers)}
        if header_name not in idx:
            return
        min_dt: datetime | None = None
        max_dt: datetime | None = None
        for row in rows:
            value = row[idx[header_name]]
            if not isinstance(value, datetime):
                continue
            if min_dt is None or value < min_dt:
                min_dt = value
            if max_dt is None or value > max_dt:
                max_dt = value
        result[min_key] = min_dt.strftime("%Y-%m") if min_dt else None
        result[max_key] = max_dt.strftime("%Y-%m") if max_dt else None

    scan_sheet("实际数据源", "日期", "actual_min", "actual_max")
    scan_sheet("预算数据源", "月份", "budget_min", "budget_max")
    return result


def infer_target_period(date_range: dict[str, str | None]) -> tuple[int | None, int | None]:
    actual_max = date_range.get("actual_max")
    budget_max = date_range.get("budget_max")
    candidates = [value for value in (actual_max, budget_max) if value]
    if not candidates:
        return None, None
    target = min(candidates) if len(candidates) == 2 else candidates[0]
    year_text, month_text = str(target).split("-")
    return int(year_text), int(month_text)


def export_period_context(run_context_dir: Path, workbook_path: Path, date_range: dict[str, str | None]) -> None:
    target_year, target_month = infer_target_period(date_range)
    payload = {
        "workbook": str(workbook_path),
        "date_range": date_range,
        "target_year": target_year,
        "target_month": target_month,
        "target_period": f"{target_year}-{target_month:02d}" if target_year and target_month else None,
        "inference_rule": "latest common covered month between actual and budget when both exist; otherwise latest available month",
    }
    (run_context_dir / "period_context.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    lines = [
        "# Period Context",
        "",
        f"- Workbook: `{workbook_path}`",
        f"- Actual coverage: `{date_range.get('actual_min')}` to `{date_range.get('actual_max')}`",
        f"- Budget coverage: `{date_range.get('budget_min')}` to `{date_range.get('budget_max')}`",
        f"- Inferred target period: `{payload['target_period']}`",
        "- Rule: use the latest common covered month when both actual and budget exist; otherwise use the latest available month.",
    ]
    (run_context_dir / "period_context.md").write_text("\n".join(lines), encoding="utf-8")


def export_category_monthly_trends(workbook_path: Path, run_root: Path, top_n: int = 10) -> None:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    if "实际数据源" not in wb.sheetnames:
        return

    ws = wb["实际数据源"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, ())
    idx = {name: i for i, name in enumerate(headers)}
    required = {"日期", "三级分类", "GMV", "销售毛利"}
    if not required.issubset(idx):
        return

    monthly_category: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"gmv": 0.0, "gp": 0.0})
    target_month_gmv: dict[str, float] = defaultdict(float)
    latest_period: str | None = None

    for row in rows:
        dt = row[idx["日期"]]
        if not isinstance(dt, datetime):
            continue
        category = normalize_text(row[idx["三级分类"]])
        period = dt.strftime("%Y-%m")
        latest_period = max(latest_period, period) if latest_period else period
        monthly_category[(period, category)]["gmv"] += float(row[idx["GMV"]] or 0)
        monthly_category[(period, category)]["gp"] += float(row[idx["销售毛利"]] or 0)

    if latest_period is None:
        return

    for (period, category), metrics in monthly_category.items():
        if period == latest_period:
            target_month_gmv[category] += metrics["gmv"]

    ranked_categories = sorted(target_month_gmv.items(), key=lambda item: item[1], reverse=True)[:top_n]
    if not ranked_categories:
        return

    total_target_month_gmv = sum(target_month_gmv.values()) or 1.0
    selected_categories = [category for category, _ in ranked_categories]
    periods = sorted({period for period, _ in monthly_category})

    structure_dir = stage_dir(run_root, "evidence") / "structure"
    structure_dir.mkdir(parents=True, exist_ok=True)
    trend_path = structure_dir / "category_monthly_trend_top10.csv"
    selection_path = structure_dir / "category_monthly_trend_selection.csv"
    note_path = structure_dir / "category_monthly_trend_selection.md"

    with trend_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "period",
                "category",
                "gmv",
                "gross_profit",
                "gross_margin",
                "target_month",
                "target_month_rank",
                "target_month_gmv",
                "target_month_gmv_share",
            ],
        )
        writer.writeheader()
        for rank, (category, category_target_gmv) in enumerate(ranked_categories, start=1):
            for period in periods:
                metrics = monthly_category.get((period, category), {"gmv": 0.0, "gp": 0.0})
                gmv = metrics["gmv"]
                gp = metrics["gp"]
                gross_margin = gp / gmv if gmv else 0.0
                writer.writerow(
                    {
                        "period": period,
                        "category": category,
                        "gmv": f"{gmv:.6f}",
                        "gross_profit": f"{gp:.6f}",
                        "gross_margin": f"{gross_margin:.8f}",
                        "target_month": latest_period,
                        "target_month_rank": rank,
                        "target_month_gmv": f"{category_target_gmv:.6f}",
                        "target_month_gmv_share": f"{category_target_gmv / total_target_month_gmv:.8f}",
                    }
                )

    selection_rows: list[dict[str, str | float | int]] = []
    for rank, (category, category_target_gmv) in enumerate(ranked_categories, start=1):
        series = [
            {
                "period": period,
                "gmv": monthly_category.get((period, category), {"gmv": 0.0})["gmv"],
                "gross_margin": (
                    monthly_category.get((period, category), {"gp": 0.0, "gmv": 0.0})["gp"]
                    / monthly_category.get((period, category), {"gp": 0.0, "gmv": 0.0})["gmv"]
                    if monthly_category.get((period, category), {"gp": 0.0, "gmv": 0.0})["gmv"]
                    else 0.0
                ),
            }
            for period in periods
        ]
        latest = series[-1]
        prev = series[-2] if len(series) >= 2 else None
        avg_margin = sum(item["gross_margin"] for item in series) / len(series)
        avg_gmv = sum(item["gmv"] for item in series) / len(series)
        mom_gmv_change = ((latest["gmv"] - prev["gmv"]) / prev["gmv"]) if prev and prev["gmv"] else 0.0
        margin_deviation = latest["gross_margin"] - avg_margin
        anomaly_score = abs(margin_deviation) * 100 + abs(mom_gmv_change) * 10
        reasons: list[str] = []
        if latest["gross_margin"] < 0:
            anomaly_score += 5
            reasons.append("latest margin negative")
        if margin_deviation <= -0.03:
            reasons.append("margin below historical average")
        elif margin_deviation >= 0.03:
            reasons.append("margin above historical average")
        if mom_gmv_change <= -0.25:
            reasons.append("sharp gmv drop")
        elif mom_gmv_change >= 0.25:
            reasons.append("sharp gmv jump")
        if not reasons:
            reasons.append("high current-month scale")
        selection_rows.append(
            {
                "category": category,
                "target_month_rank": rank,
                "target_month": latest_period,
                "target_month_gmv": category_target_gmv,
                "target_month_gmv_share": category_target_gmv / total_target_month_gmv,
                "latest_gmv": latest["gmv"],
                "latest_gross_margin": latest["gross_margin"],
                "average_gross_margin": avg_margin,
                "gmv_mom_change": mom_gmv_change,
                "anomaly_score": anomaly_score,
                "selection_reason": "; ".join(reasons),
            }
        )

    selection_rows.sort(key=lambda row: float(row["anomaly_score"]), reverse=True)
    with selection_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "category",
                "target_month_rank",
                "target_month",
                "target_month_gmv",
                "target_month_gmv_share",
                "latest_gmv",
                "latest_gross_margin",
                "average_gross_margin",
                "gmv_mom_change",
                "anomaly_score",
                "selection_reason",
            ],
        )
        writer.writeheader()
        for row in selection_rows:
            writer.writerow(row)

    top_selected = selection_rows[:3]
    lines = [
        "# Category Trend Selection",
        "",
        f"- Target month: `{latest_period}`",
        f"- Selection pool: top `{len(selected_categories)}` categories by target-month GMV",
        "- Recommendation: use 1-3 charts below in the final report, not all charts.",
        "",
    ]
    for row in top_selected:
        lines.append(
            f"- `{row['category']}` | rank=`{row['target_month_rank']}` | anomaly_score=`{float(row['anomaly_score']):.2f}` | reason=`{row['selection_reason']}`"
        )
    note_path.write_text("\n".join(lines), encoding="utf-8")


def write_profile_markdown(
    path: Path,
    workbook_path: Path,
    sheet_names: list[str],
    headers_by_sheet: dict[str, list[str]],
    confirmed_by_sheet: dict[str, list[tuple[str, str]]],
    unknown_by_sheet: dict[str, list[str]],
) -> None:
    lines = [
        "# Workbook Profile",
        "",
        f"- Workbook: `{workbook_path}`",
        f"- Sheets: {', '.join(sheet_names)}",
        "",
    ]
    for sheet_name in sheet_names:
        lines.append(f"## {sheet_name}")
        lines.append("")
        lines.append("### Headers")
        lines.append("")
        for header in headers_by_sheet[sheet_name]:
            lines.append(f"- `{header}`")
        lines.append("")
        lines.append("### Confirmed semantic mappings")
        lines.append("")
        if confirmed_by_sheet[sheet_name]:
            for source, target in confirmed_by_sheet[sheet_name]:
                lines.append(f"- `{source}` -> `{target}`")
        else:
            lines.append("- none")
        lines.append("")
        lines.append("### Fields requiring confirmation")
        lines.append("")
        if unknown_by_sheet[sheet_name]:
            for header in unknown_by_sheet[sheet_name]:
                lines.append(f"- `{header}`")
        else:
            lines.append("- none")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_mapping_draft(
    path: Path,
    confirmed_by_sheet: dict[str, list[tuple[str, str]]],
    unknown_by_sheet: dict[str, list[str]],
) -> None:
    lines = [
        "# Field Mapping Draft",
        "",
        "Use this draft before full evidence generation. Confirm uncertain fields with the user and then update `references/field_mapping_memory.md`.",
        "",
    ]
    for sheet_name in confirmed_by_sheet:
        lines.append(f"## {sheet_name}")
        lines.append("")
        lines.append("### Confirmed")
        lines.append("")
        if confirmed_by_sheet[sheet_name]:
            for source, target in confirmed_by_sheet[sheet_name]:
                lines.append(f"- `{source}` -> `{target}`")
        else:
            lines.append("- none")
        lines.append("")
        lines.append("### Need confirmation")
        lines.append("")
        if unknown_by_sheet[sheet_name]:
            for header in unknown_by_sheet[sheet_name]:
                lines.append(f"- `{header}` -> `TBD`")
        else:
            lines.append("- none")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    project_root = Path.cwd()
    skill_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description="Profile workbook headers and draft field mappings for CBEC operating analysis.")
    parser.add_argument("--input", default=str(project_root / "data" / "demo.xlsx"), help="Workbook path")
    parser.add_argument(
        "--run-root",
        default=None,
        help="Root directory for this analysis run. If omitted, a new timestamped run directory is created under output/runs/.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Workbook evidence output directory. Defaults to <run_root>/evidence/workbook/",
    )
    parser.add_argument("--mapping-memory", default=str(skill_root / "references" / "field_mapping_memory.md"), help="Mapping memory markdown path")
    args = parser.parse_args()

    workbook_path = Path(args.input)
    run_root = resolve_run_root(project_root, args.run_root)
    output_dir = Path(args.output_dir) if args.output_dir else stage_dir(run_root, "evidence") / "workbook"
    run_context_dir = stage_dir(run_root, "run_context")
    mapping_path = Path(args.mapping_memory)
    output_dir.mkdir(parents=True, exist_ok=True)

    mappings = load_mapping_memory(mapping_path)
    sheet_names, headers_by_sheet, confirmed_by_sheet, unknown_by_sheet = workbook_profile(workbook_path, mappings)
    date_range = sheet_date_range(workbook_path)

    write_profile_markdown(
        output_dir / "workbook_profile.md",
        workbook_path,
        sheet_names,
        headers_by_sheet,
        confirmed_by_sheet,
        unknown_by_sheet,
    )
    write_mapping_draft(output_dir / "field_mapping_draft.md", confirmed_by_sheet, unknown_by_sheet)
    payload = {
        "workbook": str(workbook_path),
        "sheets": sheet_names,
        "headers_by_sheet": headers_by_sheet,
        "confirmed_by_sheet": {k: [{"source": s, "target": t} for s, t in v] for k, v in confirmed_by_sheet.items()},
        "unknown_by_sheet": unknown_by_sheet,
    }
    (output_dir / "header_index.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    export_period_context(run_context_dir, workbook_path, date_range)
    export_category_monthly_trends(workbook_path, run_root, top_n=10)
    print(run_root)


if __name__ == "__main__":
    main()
