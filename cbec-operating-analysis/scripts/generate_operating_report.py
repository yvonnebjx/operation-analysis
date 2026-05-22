from __future__ import annotations

import argparse
from collections import defaultdict
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Callable

import openpyxl
from run_paths import resolve_run_root, stage_dir


TARGET_YEAR = 2024
TARGET_MONTH = 11

PERIODS = ("mtd", "ytd", "nov23", "ytd23")
GROUPS = ("channel", "region", "product_line", "category", "grade")
IGNORED_MODELS = {"", "未知", "毛保费用"}


def sheet_max_datetime(ws, header_name: str) -> datetime | None:
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    idx = {name: i for i, name in enumerate(headers)}
    if header_name not in idx:
        return None
    max_dt: datetime | None = None
    for row in rows:
        value = row[idx[header_name]]
        if isinstance(value, datetime):
            if max_dt is None or value > max_dt:
                max_dt = value
    return max_dt


def detect_target_period(workbook_path: Path) -> tuple[int, int]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    actual_max = sheet_max_datetime(wb["实际数据源"], "日期") if "实际数据源" in wb.sheetnames else None
    budget_max = sheet_max_datetime(wb["预算数据源"], "月份") if "预算数据源" in wb.sheetnames else None

    target = actual_max or budget_max
    if actual_max and budget_max:
        target = min(actual_max, budget_max)
    if target is None:
        return TARGET_YEAR, TARGET_MONTH
    return target.year, target.month


def period_context() -> dict[str, str | int]:
    last_year = TARGET_YEAR - 1
    return {
        "target_year": TARGET_YEAR,
        "target_month": TARGET_MONTH,
        "last_year": last_year,
        "month_label": f"{TARGET_MONTH}月",
        "current_month_text": f"{TARGET_YEAR}年{TARGET_MONTH}月",
        "last_year_month_text": f"{last_year}年{TARGET_MONTH}月",
        "current_ytd_text": f"{TARGET_YEAR}年1-{TARGET_MONTH}月",
        "last_ytd_text": f"{last_year}年1-{TARGET_MONTH}月",
    }


def new_metrics() -> dict[str, float]:
    return {
        "qty": 0.0,
        "gmv": 0.0,
        "discount": 0.0,
        "vat": 0.0,
        "refund": 0.0,
        "revenue": 0.0,
        "purchase": 0.0,
        "headship": 0.0,
        "ship": 0.0,
        "ads": 0.0,
        "commission": 0.0,
        "fba": 0.0,
        "storage": 0.0,
        "overseas": 0.0,
        "other": 0.0,
        "gp": 0.0,
    }


def clone_metrics(metrics: dict[str, float]) -> dict[str, float]:
    return deepcopy(metrics)


def add_metrics(dest: dict[str, float], src: dict[str, float]) -> None:
    for key, value in src.items():
        if value in (None, "", "-", " - "):
            numeric = 0.0
        else:
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                text = str(value).replace(",", "").strip()
                numeric = float(text) if text not in ("", "-", "- ") else 0.0
        dest[key] += numeric


def safe_div(num: float, den: float) -> float:
    return num / den if den else 0.0


def normalize_text(value: object, default: str = "未知") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def normalize_region(value: object) -> str:
    return normalize_text(value).upper()


def rates(metrics: dict[str, float]) -> dict[str, float]:
    gmv = metrics["gmv"]
    logistics_total = metrics["headship"] + metrics["overseas"]
    platform_fee = (
        metrics["ads"]
        + metrics["commission"]
        + metrics["fba"]
        + metrics["storage"]
        + metrics["other"]
    )
    return {
        "gross_margin": safe_div(metrics["gp"], gmv),
        "discount_rate": safe_div(metrics["discount"], gmv),
        "vat_rate": safe_div(metrics["vat"], gmv),
        "refund_rate": safe_div(metrics["refund"], gmv),
        "purchase_rate": safe_div(metrics["purchase"], gmv),
        "logistics_rate": safe_div(logistics_total, gmv),
        "headship_rate": safe_div(metrics["headship"], gmv),
        "overseas_rate": safe_div(metrics["overseas"], gmv),
        "ads_rate": safe_div(metrics["ads"], gmv),
        "commission_rate": safe_div(metrics["commission"], gmv),
        "delivery_rate": safe_div(metrics["fba"], gmv),
        "storage_rate": safe_div(metrics["storage"], gmv),
        "other_rate": safe_div(metrics["other"], gmv),
        "platform_fee_rate": safe_div(platform_fee, gmv),
        "revenue_rate": safe_div(metrics["revenue"], gmv),
    }


def money_m(value: float) -> str:
    return f"{value / 1_000_000:.2f}"


def money(value: float) -> str:
    return f"{value:.2f}"


def qty_k(value: float) -> str:
    return f"{value / 1_000:.1f}"


def pct(value: float) -> str:
    return f"{value * 100:.2f}%"


def pp(value: float) -> str:
    return f"{value * 100:.2f}pp"


def pp_change_text(value: float) -> str:
    direction = "上升" if value >= 0 else "下降"
    return f"{direction} {abs(value) * 100:.2f}pp"


def ratio(actual: float, budget: float) -> str:
    return f"{safe_div(actual, budget) * 100:.1f}%"


def usd_unit(value: float) -> str:
    return f"{value:.2f}"


def delta_m(value: float) -> str:
    return f"{value / 1_000_000:+.2f}"


def delta_k(value: float) -> str:
    return f"{value / 1_000:+.1f}"


def signed_pct(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value * 100:+.1f}%"


def signed_pp(value: float) -> str:
    return f"{value * 100:+.2f}pp"


def yoy_change(current: float, previous: float) -> float | None:
    if previous == 0:
        return None
    delta = current - previous
    if previous > 0:
        return delta / previous
    return 1 - (delta / previous)


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def unit_amount(metrics: dict[str, float], key: str) -> float:
    return safe_div(metrics[key], metrics["qty"])


def gmv_price_volume_analysis(actual: dict[str, float], budget: dict[str, float]) -> dict[str, float]:
    actual_qty = actual["qty"]
    budget_qty = budget["qty"]
    actual_asp = unit_amount(actual, "gmv")
    budget_asp = unit_amount(budget, "gmv")
    volume_effect = (actual_qty - budget_qty) * budget_asp
    price_effect = actual_qty * (actual_asp - budget_asp)
    return {
        "volume_effect": volume_effect,
        "price_effect": price_effect,
        "total_effect": actual["gmv"] - budget["gmv"],
        "actual_asp": actual_asp,
        "budget_asp": budget_asp,
        "actual_qty": actual_qty,
        "budget_qty": budget_qty,
    }


def gmv_pv_table(actual: dict[str, float], budget: dict[str, float]) -> str:
    analysis = gmv_price_volume_analysis(actual, budget)
    rows = [
        ["预算销量(千件)", qty_k(analysis["budget_qty"]), "-", "-"],
        ["实际销量(千件)", qty_k(analysis["actual_qty"]), "-", "-"],
        ["预算GMV单价(USD/件)", usd_unit(analysis["budget_asp"]), "-", "-"],
        ["实际GMV单价(USD/件)", usd_unit(analysis["actual_asp"]), "-", "-"],
        ["量差影响", "按预算GMV单价测算", money_m(analysis["volume_effect"]), pct(safe_div(analysis["volume_effect"], analysis["total_effect"])) if analysis["total_effect"] else "-"],
        ["价差影响", "按实际销量测算单价差", money_m(analysis["price_effect"]), pct(safe_div(analysis["price_effect"], analysis["total_effect"])) if analysis["total_effect"] else "-"],
        ["GMV总差异", "实际GMV-预算GMV", money_m(analysis["total_effect"]), "100.00%"],
    ]
    return markdown_table(["项目", "说明", "影响(USD M)", "占GMV差异"], rows)


def gross_profit_bridge(actual: dict[str, float], budget: dict[str, float]) -> dict[str, object]:
    actual_qty = actual["qty"]
    budget_qty = budget["qty"]
    budget_gp_per_unit = unit_amount(budget, "gp")

    factor_defs = [
        ("gmv", "销售单价/结构", 1),
        ("discount", "折扣单价", -1),
        ("vat", "VAT单价", -1),
        ("refund", "退款单价", -1),
        ("purchase", "采购单价", -1),
        ("headship", "头程单价(含二程/关税)", -1),
        ("overseas", "海外仓单价", -1),
        ("ads", "广告单价", -1),
        ("commission", "佣金单价", -1),
        ("fba", "配送费单价", -1),
        ("storage", "仓储单价", -1),
        ("other", "其他费用单价", -1),
    ]

    rows: list[dict[str, object]] = []
    volume_effect = (actual_qty - budget_qty) * budget_gp_per_unit
    rows.append(
        {
            "factor": "销量差异",
            "budget_unit": budget_qty,
            "actual_unit": actual_qty,
            "unit_diff": actual_qty - budget_qty,
            "impact": volume_effect,
            "display": "qty",
        }
    )

    for key, label, sign in factor_defs:
        budget_unit = unit_amount(budget, key)
        actual_unit = unit_amount(actual, key)
        impact = actual_qty * (actual_unit - budget_unit) * sign
        rows.append(
            {
                "factor": label,
                "budget_unit": budget_unit,
                "actual_unit": actual_unit,
                "unit_diff": actual_unit - budget_unit,
                "impact": impact,
                "display": "usd",
            }
        )

    total_effect = actual["gp"] - budget["gp"]
    unit_economics_effect = total_effect - volume_effect
    return {
        "rows": rows,
        "volume_effect": volume_effect,
        "unit_economics_effect": unit_economics_effect,
        "total_effect": total_effect,
        "actual_qty": actual_qty,
        "budget_qty": budget_qty,
    }


def gross_profit_bridge_table(actual: dict[str, float], budget: dict[str, float]) -> str:
    bridge = gross_profit_bridge(actual, budget)
    rows = []
    for row in bridge["rows"]:
        if row["display"] == "qty":
            budget_unit = qty_k(row["budget_unit"])
            actual_unit = qty_k(row["actual_unit"])
            unit_diff = qty_k(row["unit_diff"])
        else:
            budget_unit = usd_unit(row["budget_unit"])
            actual_unit = usd_unit(row["actual_unit"])
            unit_diff = usd_unit(row["unit_diff"])
        rows.append(
            [
                row["factor"],
                budget_unit,
                actual_unit,
                unit_diff,
                money_m(row["impact"]),
            ]
        )
    rows.append(["合计利润差异", "-", "-", "-", money_m(bridge["total_effect"])])
    return markdown_table(
        ["因素", "预算单件/预算销量", "实际单件/实际销量", "差异", "对毛利影响(USD M)"],
        rows,
    )


def summarize_bridge(bridge: dict[str, object], top_n: int = 4) -> tuple[list[str], list[str]]:
    rows = [row for row in bridge["rows"] if row["factor"] != "销量差异"]
    adverse = sorted((row for row in rows if row["impact"] < 0), key=lambda row: row["impact"])
    favorable = sorted((row for row in rows if row["impact"] > 0), key=lambda row: row["impact"], reverse=True)
    adverse_text = [f"{row['factor']} {money_m(row['impact'])}M" for row in adverse[:top_n]]
    favorable_text = [f"{row['factor']} +{money_m(row['impact'])}M" for row in favorable[:top_n]]
    return adverse_text, favorable_text


def channel_profit_variance_table(
    actual_groups: dict[str, dict[str, float]],
    budget_groups: dict[str, dict[str, float]],
    limit: int = 8,
) -> str:
    keys = sorted(
        actual_groups.keys() | budget_groups.keys(),
        key=lambda key: actual_groups.get(key, new_metrics())["gp"] - budget_groups.get(key, new_metrics())["gp"],
    )
    rows = []
    for key in keys[:limit]:
        actual = actual_groups.get(key, new_metrics())
        budget = budget_groups.get(key, new_metrics())
        bridge = gross_profit_bridge(actual, budget)
        adverse, _ = summarize_bridge(bridge, top_n=2)
        rows.append(
            [
                key,
                money_m(actual["gmv"]),
                money_m(actual["gp"] - budget["gp"]),
                money_m(bridge["volume_effect"]),
                money_m(bridge["unit_economics_effect"]),
                "；".join(adverse) if adverse else "-",
            ]
        )
    return markdown_table(
        ["渠道", "实际GMV", "毛利差异", "量差影响", "单位经济差影响", "主要利空因素"],
        rows,
    )


def aggregate_composite_groups(
    analyzer: Analyzer,
    source: str,
    period: str,
    dimensions: tuple[str, ...],
) -> dict[str, dict[str, float]]:
    grouped: defaultdict[str, dict[str, float]] = defaultdict(new_metrics)
    for combo_key, metrics in analyzer.combos[(source, period)].items():
        info = analyzer.combo_info.get(combo_key, {})
        label_parts = [info.get(dimension, "未知") for dimension in dimensions]
        label = " / ".join(label_parts)
        add_metrics(grouped[label], metrics)
    return dict(grouped)


def aggregate_filtered_groups(
    analyzer: Analyzer,
    source: str,
    period: str,
    group_key: str,
    filters: dict[str, str],
) -> dict[str, dict[str, float]]:
    grouped: defaultdict[str, dict[str, float]] = defaultdict(new_metrics)
    for combo_key, metrics in analyzer.combos[(source, period)].items():
        info = analyzer.combo_info.get(combo_key, {})
        if any(info.get(field, "未知") != value for field, value in filters.items()):
            continue
        label = info.get(group_key, "未知")
        add_metrics(grouped[label], metrics)
    return dict(grouped)


def composite_profit_variance_table(
    actual_groups: dict[str, dict[str, float]],
    budget_groups: dict[str, dict[str, float]],
    title: str,
    limit: int = 12,
) -> str:
    keys = sorted(
        actual_groups.keys() | budget_groups.keys(),
        key=lambda key: actual_groups.get(key, new_metrics())["gp"] - budget_groups.get(key, new_metrics())["gp"],
    )
    rows = []
    for key in keys[:limit]:
        actual = actual_groups.get(key, new_metrics())
        budget = budget_groups.get(key, new_metrics())
        bridge = gross_profit_bridge(actual, budget)
        adverse, _ = summarize_bridge(bridge, top_n=2)
        rows.append(
            [
                key,
                money_m(actual["gmv"]),
                money_m(actual["gp"] - budget["gp"]),
                money_m(bridge["volume_effect"]),
                money_m(bridge["unit_economics_effect"]),
                "；".join(adverse) if adverse else "-",
            ]
        )
    return "\n".join(
        [
            f"### {title}",
            "",
            markdown_table(
                ["组合", "实际GMV", "毛利差异", "量差影响", "单位经济差影响", "主要利空因素"],
                rows,
            ),
        ]
    )


def channel_refund_sections(analyzer: Analyzer, limit_category: int = 8, limit_region: int = 6) -> list[str]:
    lines: list[str] = []
    channel_groups = analyzer.groups[("actual", "mtd", "channel")]
    channels = [
        channel
        for channel, metrics in sorted(channel_groups.items(), key=lambda item: item[1]["gmv"], reverse=True)
        if metrics["gmv"] > 0
    ]
    for channel in channels:
        actual_channel = channel_groups[channel]
        last_year_channel = analyzer.groups[("actual", "nov23", "channel")].get(channel, new_metrics())
        category_actual = aggregate_filtered_groups(analyzer, "actual", "mtd", "category", {"channel": channel})
        category_last = aggregate_filtered_groups(analyzer, "actual", "nov23", "category", {"channel": channel})
        region_actual = aggregate_filtered_groups(analyzer, "actual", "mtd", "region", {"channel": channel})
        region_last = aggregate_filtered_groups(analyzer, "actual", "nov23", "region", {"channel": channel})

        lines.append(f"### 渠道：{channel}")
        lines.append("")
        lines.append(
            f"11月该渠道 GMV {money_m(actual_channel['gmv'])}M，退款率 {pct(rates(actual_channel)['refund_rate'])}，较 2023 年 11 月变动 "
            + (
                signed_pp(rates(actual_channel)["refund_rate"] - rates(last_year_channel)["refund_rate"])
                if last_year_channel["gmv"]
                else "-"
            )
            + "。"
        )
        lines.append("")
        lines.append("分类退款率")
        lines.append("")
        lines.append(refund_table(category_actual, category_last, limit=limit_category, min_gmv=100_000))
        lines.append("")
        lines.append("区域退款率")
        lines.append("")
        lines.append(refund_table(region_actual, region_last, limit=limit_region, min_gmv=0))
        lines.append("")
    return lines


class Analyzer:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        self.totals: dict[tuple[str, str], dict[str, float]] = {}
        self.groups: dict[tuple[str, str, str], defaultdict[str, dict[str, float]]] = {}
        self.models: dict[tuple[str, str], defaultdict[str, dict[str, float]]] = {}
        self.combos: dict[tuple[str, str], defaultdict[tuple[str, str, str], dict[str, float]]] = {}
        self.model_info: dict[str, dict[str, str]] = {}
        self.combo_info: dict[tuple[str, str, str], dict[str, str]] = {}
        for source in ("actual", "budget"):
            for period in PERIODS:
                self.totals[(source, period)] = new_metrics()
                self.models[(source, period)] = defaultdict(new_metrics)
                self.combos[(source, period)] = defaultdict(new_metrics)
                for group in GROUPS:
                    self.groups[(source, period, group)] = defaultdict(new_metrics)

    def update_info(self, model: str, info: dict[str, str]) -> None:
        current = self.model_info.setdefault(model, {})
        for key, value in info.items():
            if value and value != "未知" and not current.get(key):
                current[key] = value

    def update_combo_info(self, combo_key: tuple[str, str, str], info: dict[str, str]) -> None:
        current = self.combo_info.setdefault(combo_key, {})
        for key, value in info.items():
            if value and value != "未知" and not current.get(key):
                current[key] = value

    def apply_row(
        self,
        source: str,
        period: str,
        metrics: dict[str, float],
        info: dict[str, str],
    ) -> None:
        add_metrics(self.totals[(source, period)], metrics)
        for group in GROUPS:
            add_metrics(self.groups[(source, period, group)][info[group]], metrics)
        add_metrics(self.models[(source, period)][info["model"]], metrics)
        combo_key = (info["channel"], info["region"], info["model"])
        add_metrics(self.combos[(source, period)][combo_key], metrics)
        self.update_info(info["model"], info)
        self.update_combo_info(combo_key, info)

    def read_budget(self) -> None:
        ws = self.wb["预算数据源"]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows)
        idx = {name: i for i, name in enumerate(headers)}
        for row in rows:
            dt = row[idx["月份"]]
            if not isinstance(dt, datetime):
                continue
            if dt.year != TARGET_YEAR or dt.month > TARGET_MONTH:
                continue
            info = {
                "channel": normalize_text(row[idx["渠道"]]),
                "region": normalize_region(row[idx["地区"]]),
                "product_line": normalize_text(row[idx["大品类"]]),
                "category": normalize_text(row[idx["新三级类目"]]),
                "grade": normalize_text(row[idx["类目等级"]]),
                "model": normalize_text(row[idx["产品型号"]]),
                "brand": normalize_text(row[idx["品牌"]]),
            }
            metrics = {
                "qty": row[idx["预算销售数量"]] or 0,
                "gmv": row[idx["预算GMV"]] or 0,
                "discount": row[idx["预算折扣金额"]] or 0,
                "vat": row[idx["预算VAT"]] or 0,
                "refund": row[idx["预算退款金额"]] or 0,
                "revenue": row[idx["预算营业收入"]] or 0,
                "purchase": row[idx["预算采购金额"]] or 0,
                "headship": (
                    (row[idx["预算头程成本"]] or 0)
                    + (row[idx["预算关税"]] or 0)
                    + (row[idx["预算二程配送费"]] or 0)
                ),
                "ship": (
                    (row[idx["预算头程成本"]] or 0)
                    + (row[idx["预算关税"]] or 0)
                    + (row[idx["预算二程配送费"]] or 0)
                    + (row[idx["预算海外仓费用"]] or 0)
                ),
                "ads": row[idx["预算广告费用"]] or 0,
                "commission": row[idx["预算平台佣金"]] or 0,
                "fba": (row[idx["预算FBA配送费"]] or 0) + (row[idx["预算FBM配送费"]] or 0),
                "storage": row[idx["预算平台仓储费"]] or 0,
                "overseas": row[idx["预算海外仓费用"]] or 0,
                "other": (row[idx["预算平台其他费用"]] or 0) + (row[idx["预算异常费用"]] or 0),
                "gp": row[idx["预算运营毛利"]] or 0,
            }
            self.apply_row("budget", "ytd", metrics, info)
            if dt.month == TARGET_MONTH:
                self.apply_row("budget", "mtd", metrics, info)

    def read_actual(self) -> None:
        ws = self.wb["实际数据源"]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows)
        idx = {name: i for i, name in enumerate(headers)}
        for row in rows:
            dt = row[idx["日期"]]
            if not isinstance(dt, datetime):
                continue
            info = {
                "channel": normalize_text(row[idx["渠道"]]),
                "region": normalize_region(row[idx["地区"]]),
                "product_line": normalize_text(row[idx["产品线"]]),
                "category": normalize_text(row[idx["三级分类"]]),
                "grade": normalize_text(row[idx["类目等级"]]),
                "model": normalize_text(row[idx["MODEL"]]),
                "brand": normalize_text(row[idx["品牌"]]),
            }
            metrics = {
                "qty": row[idx["销售数量"]] or 0,
                "gmv": row[idx["GMV"]] or 0,
                "discount": row[idx["折扣金额合计"]] or 0,
                "vat": row[idx["VAT费用合计"]] or 0,
                "refund": row[idx["退货额合计"]] or 0,
                "revenue": row[idx["销售收入"]] or 0,
                "purchase": row[idx["采购成本合计"]] or 0,
                "headship": row[idx["运费成本合计"]] or 0,
                "ship": (row[idx["运费成本合计"]] or 0) + (row[idx["海外仓费用"]] or 0),
                "ads": row[idx["广告费用合计"]] or 0,
                "commission": row[idx["订单佣金"]] or 0,
                "fba": row[idx["订单FBA费用"]] or 0,
                "storage": row[idx["平台仓储费合计"]] or 0,
                "overseas": row[idx["海外仓费用"]] or 0,
                "other": row[idx["其他费用合计"]] or 0,
                "gp": row[idx["销售毛利"]] or 0,
            }
            if dt.year == TARGET_YEAR and dt.month <= TARGET_MONTH:
                self.apply_row("actual", "ytd", metrics, info)
                if dt.month == TARGET_MONTH:
                    self.apply_row("actual", "mtd", metrics, info)
            elif dt.year == TARGET_YEAR - 1 and dt.month <= TARGET_MONTH:
                self.apply_row("actual", "ytd23", metrics, info)
                if dt.month == TARGET_MONTH:
                    self.apply_row("actual", "nov23", metrics, info)

    def run(self) -> None:
        self.read_budget()
        self.read_actual()

    def matched_gmv_share(self, period: str) -> tuple[float, float]:
        actual = self.combos[("actual", period)]
        budget = self.combos[("budget", period)]
        total_actual_gmv = sum(metrics["gmv"] for metrics in actual.values())
        matched_actual_gmv = sum(
            metrics["gmv"] for combo, metrics in actual.items() if combo in budget
        )
        return safe_div(matched_actual_gmv, total_actual_gmv), total_actual_gmv - matched_actual_gmv


def metrics_table(actual: dict[str, float], budget: dict[str, float]) -> str:
    actual_rates = rates(actual)
    budget_rates = rates(budget)
    rows = [
        ["销售数量(千件)", qty_k(actual["qty"]), qty_k(budget["qty"]), qty_k(actual["qty"] - budget["qty"]), ratio(actual["qty"], budget["qty"])],
        ["GMV(USD M)", money_m(actual["gmv"]), money_m(budget["gmv"]), money_m(actual["gmv"] - budget["gmv"]), ratio(actual["gmv"], budget["gmv"])],
        ["营业收入(USD M)", money_m(actual["revenue"]), money_m(budget["revenue"]), money_m(actual["revenue"] - budget["revenue"]), ratio(actual["revenue"], budget["revenue"])],
        ["销售毛利(USD M)", money_m(actual["gp"]), money_m(budget["gp"]), money_m(actual["gp"] - budget["gp"]), ratio(actual["gp"], budget["gp"])],
        ["毛利率", pct(actual_rates["gross_margin"]), pct(budget_rates["gross_margin"]), pp(actual_rates["gross_margin"] - budget_rates["gross_margin"]), "-"],
        ["折扣率", pct(actual_rates["discount_rate"]), pct(budget_rates["discount_rate"]), pp(actual_rates["discount_rate"] - budget_rates["discount_rate"]), "-"],
        ["VAT费率", pct(actual_rates["vat_rate"]), pct(budget_rates["vat_rate"]), pp(actual_rates["vat_rate"] - budget_rates["vat_rate"]), "-"],
        ["退款率", pct(actual_rates["refund_rate"]), pct(budget_rates["refund_rate"]), pp(actual_rates["refund_rate"] - budget_rates["refund_rate"]), "-"],
        ["采购费率", pct(actual_rates["purchase_rate"]), pct(budget_rates["purchase_rate"]), pp(actual_rates["purchase_rate"] - budget_rates["purchase_rate"]), "-"],
        ["物流费率", pct(actual_rates["logistics_rate"]), pct(budget_rates["logistics_rate"]), pp(actual_rates["logistics_rate"] - budget_rates["logistics_rate"]), "-"],
        ["广告费率", pct(actual_rates["ads_rate"]), pct(budget_rates["ads_rate"]), pp(actual_rates["ads_rate"] - budget_rates["ads_rate"]), "-"],
        ["平台费率", pct(actual_rates["platform_fee_rate"]), pct(budget_rates["platform_fee_rate"]), pp(actual_rates["platform_fee_rate"] - budget_rates["platform_fee_rate"]), "-"],
    ]
    return markdown_table(["指标", "实际", "预算", "差异", "达成率"], rows)


def combined_metrics_table(
    actual: dict[str, float], budget: dict[str, float], last_year: dict[str, float]
) -> str:
    actual_rates = rates(actual)
    budget_rates = rates(budget)
    last_year_rates = rates(last_year)
    rows = [
        [
            "销售数量(千件)",
            qty_k(actual["qty"]),
            qty_k(budget["qty"]),
            delta_k(actual["qty"] - budget["qty"]),
            ratio(actual["qty"], budget["qty"]),
            qty_k(last_year["qty"]),
            delta_k(actual["qty"] - last_year["qty"]),
            signed_pct(yoy_change(actual["qty"], last_year["qty"])),
        ],
        [
            "GMV(USD M)",
            money_m(actual["gmv"]),
            money_m(budget["gmv"]),
            delta_m(actual["gmv"] - budget["gmv"]),
            ratio(actual["gmv"], budget["gmv"]),
            money_m(last_year["gmv"]),
            delta_m(actual["gmv"] - last_year["gmv"]),
            signed_pct(yoy_change(actual["gmv"], last_year["gmv"])),
        ],
        [
            "营业收入(USD M)",
            money_m(actual["revenue"]),
            money_m(budget["revenue"]),
            delta_m(actual["revenue"] - budget["revenue"]),
            ratio(actual["revenue"], budget["revenue"]),
            money_m(last_year["revenue"]),
            delta_m(actual["revenue"] - last_year["revenue"]),
            signed_pct(yoy_change(actual["revenue"], last_year["revenue"])),
        ],
        [
            "销售毛利(USD M)",
            money_m(actual["gp"]),
            money_m(budget["gp"]),
            delta_m(actual["gp"] - budget["gp"]),
            ratio(actual["gp"], budget["gp"]),
            money_m(last_year["gp"]),
            delta_m(actual["gp"] - last_year["gp"]),
            signed_pct(yoy_change(actual["gp"], last_year["gp"])),
        ],
        [
            "毛利率",
            pct(actual_rates["gross_margin"]),
            pct(budget_rates["gross_margin"]),
            signed_pp(actual_rates["gross_margin"] - budget_rates["gross_margin"]),
            "-",
            pct(last_year_rates["gross_margin"]),
            signed_pp(actual_rates["gross_margin"] - last_year_rates["gross_margin"]),
            "-",
        ],
        [
            "折扣率",
            pct(actual_rates["discount_rate"]),
            pct(budget_rates["discount_rate"]),
            signed_pp(actual_rates["discount_rate"] - budget_rates["discount_rate"]),
            "-",
            pct(last_year_rates["discount_rate"]),
            signed_pp(actual_rates["discount_rate"] - last_year_rates["discount_rate"]),
            "-",
        ],
        [
            "VAT费率",
            pct(actual_rates["vat_rate"]),
            pct(budget_rates["vat_rate"]),
            signed_pp(actual_rates["vat_rate"] - budget_rates["vat_rate"]),
            "-",
            pct(last_year_rates["vat_rate"]),
            signed_pp(actual_rates["vat_rate"] - last_year_rates["vat_rate"]),
            "-",
        ],
        [
            "退款率",
            pct(actual_rates["refund_rate"]),
            pct(budget_rates["refund_rate"]),
            signed_pp(actual_rates["refund_rate"] - budget_rates["refund_rate"]),
            "-",
            pct(last_year_rates["refund_rate"]),
            signed_pp(actual_rates["refund_rate"] - last_year_rates["refund_rate"]),
            "-",
        ],
        [
            "采购费率",
            pct(actual_rates["purchase_rate"]),
            pct(budget_rates["purchase_rate"]),
            signed_pp(actual_rates["purchase_rate"] - budget_rates["purchase_rate"]),
            "-",
            pct(last_year_rates["purchase_rate"]),
            signed_pp(actual_rates["purchase_rate"] - last_year_rates["purchase_rate"]),
            "-",
        ],
        [
            "头程(含二程/关税)",
            pct(actual_rates["headship_rate"]),
            pct(budget_rates["headship_rate"]),
            signed_pp(actual_rates["headship_rate"] - budget_rates["headship_rate"]),
            "-",
            pct(last_year_rates["headship_rate"]),
            signed_pp(actual_rates["headship_rate"] - last_year_rates["headship_rate"]),
            "-",
        ],
        [
            "海外仓",
            pct(actual_rates["overseas_rate"]),
            pct(budget_rates["overseas_rate"]),
            signed_pp(actual_rates["overseas_rate"] - budget_rates["overseas_rate"]),
            "-",
            pct(last_year_rates["overseas_rate"]),
            signed_pp(actual_rates["overseas_rate"] - last_year_rates["overseas_rate"]),
            "-",
        ],
        [
            "广告费率",
            pct(actual_rates["ads_rate"]),
            pct(budget_rates["ads_rate"]),
            signed_pp(actual_rates["ads_rate"] - budget_rates["ads_rate"]),
            "-",
            pct(last_year_rates["ads_rate"]),
            signed_pp(actual_rates["ads_rate"] - last_year_rates["ads_rate"]),
            "-",
        ],
        [
            "佣金",
            pct(actual_rates["commission_rate"]),
            pct(budget_rates["commission_rate"]),
            signed_pp(actual_rates["commission_rate"] - budget_rates["commission_rate"]),
            "-",
            pct(last_year_rates["commission_rate"]),
            signed_pp(actual_rates["commission_rate"] - last_year_rates["commission_rate"]),
            "-",
        ],
        [
            "配送费",
            pct(actual_rates["delivery_rate"]),
            pct(budget_rates["delivery_rate"]),
            signed_pp(actual_rates["delivery_rate"] - budget_rates["delivery_rate"]),
            "-",
            pct(last_year_rates["delivery_rate"]),
            signed_pp(actual_rates["delivery_rate"] - last_year_rates["delivery_rate"]),
            "-",
        ],
        [
            "平台仓储费",
            pct(actual_rates["storage_rate"]),
            pct(budget_rates["storage_rate"]),
            signed_pp(actual_rates["storage_rate"] - budget_rates["storage_rate"]),
            "-",
            pct(last_year_rates["storage_rate"]),
            signed_pp(actual_rates["storage_rate"] - last_year_rates["storage_rate"]),
            "-",
        ],
        [
            "平台其他费用",
            pct(actual_rates["other_rate"]),
            pct(budget_rates["other_rate"]),
            signed_pp(actual_rates["other_rate"] - budget_rates["other_rate"]),
            "-",
            pct(last_year_rates["other_rate"]),
            signed_pp(actual_rates["other_rate"] - last_year_rates["other_rate"]),
            "-",
        ],
    ]
    return markdown_table(
        ["指标", "实际", "预算", "较预算", "达成率", "去年同期", "同比差异", "同比"],
        rows,
    )


def yoy_rate_table(current: dict[str, float], last_year: dict[str, float]) -> str:
    current_rates = rates(current)
    last_year_rates = rates(last_year)
    ctx = period_context()
    rows = [
        ["GMV(USD M)", money_m(current["gmv"]), money_m(last_year["gmv"]), money_m(current["gmv"] - last_year["gmv"]), ratio(current["gmv"], last_year["gmv"])],
        ["销售毛利(USD M)", money_m(current["gp"]), money_m(last_year["gp"]), money_m(current["gp"] - last_year["gp"]), ratio(current["gp"], last_year["gp"])],
        ["毛利率", pct(current_rates["gross_margin"]), pct(last_year_rates["gross_margin"]), pp(current_rates["gross_margin"] - last_year_rates["gross_margin"]), "-"],
        ["折扣率", pct(current_rates["discount_rate"]), pct(last_year_rates["discount_rate"]), pp(current_rates["discount_rate"] - last_year_rates["discount_rate"]), "-"],
        ["退款率", pct(current_rates["refund_rate"]), pct(last_year_rates["refund_rate"]), pp(current_rates["refund_rate"] - last_year_rates["refund_rate"]), "-"],
        ["采购费率", pct(current_rates["purchase_rate"]), pct(last_year_rates["purchase_rate"]), pp(current_rates["purchase_rate"] - last_year_rates["purchase_rate"]), "-"],
        ["物流费率", pct(current_rates["logistics_rate"]), pct(last_year_rates["logistics_rate"]), pp(current_rates["logistics_rate"] - last_year_rates["logistics_rate"]), "-"],
        ["广告费率", pct(current_rates["ads_rate"]), pct(last_year_rates["ads_rate"]), pp(current_rates["ads_rate"] - last_year_rates["ads_rate"]), "-"],
        ["平台费率", pct(current_rates["platform_fee_rate"]), pct(last_year_rates["platform_fee_rate"]), pp(current_rates["platform_fee_rate"] - last_year_rates["platform_fee_rate"]), "-"],
    ]
    return markdown_table(["指标", f"{ctx['current_month_text']}实际", f"{ctx['last_year_month_text']}实际", "同比差异", "同比"], rows)


def safe_min_metric(
    items: dict[str, dict[str, float]],
    min_gmv: float,
    metric_key: str,
) -> tuple[str, dict[str, float]] | None:
    candidates = [(key, metrics) for key, metrics in items.items() if metrics["gmv"] >= min_gmv]
    if not candidates:
        return None
    return min(candidates, key=lambda item: rates(item[1])[metric_key])


def safe_max_metric(
    items: dict[str, dict[str, float]],
    min_gmv: float,
    metric_key: str,
) -> tuple[str, dict[str, float]] | None:
    candidates = [(key, metrics) for key, metrics in items.items() if metrics["gmv"] >= min_gmv]
    if not candidates:
        return None
    return max(candidates, key=lambda item: rates(item[1])[metric_key])


def top_groups_table(
    actual_groups: dict[str, dict[str, float]],
    budget_groups: dict[str, dict[str, float]],
    last_year_groups: dict[str, dict[str, float]],
    total_actual: dict[str, float],
    total_budget: dict[str, float],
    total_last_year: dict[str, float],
    limit: int = 8,
) -> str:
    keys = sorted(
        actual_groups.keys() | budget_groups.keys() | last_year_groups.keys(),
        key=lambda key: actual_groups.get(key, new_metrics())["gmv"],
        reverse=True,
    )
    rows: list[list[str]] = []
    total_gp_actual = total_actual["gp"]
    total_gp_budget = total_budget["gp"]
    total_gp_last_year = total_last_year["gp"]
    for key in keys[:limit]:
        actual = actual_groups.get(key, new_metrics())
        budget = budget_groups.get(key, new_metrics())
        last_year = last_year_groups.get(key, new_metrics())
        actual_gmv_share = safe_div(actual["gmv"], total_actual["gmv"])
        budget_gmv_share = safe_div(budget["gmv"], total_budget["gmv"])
        last_year_gmv_share = safe_div(last_year["gmv"], total_last_year["gmv"])
        actual_gp_share = safe_div(actual["gp"], total_gp_actual)
        budget_gp_share = safe_div(budget["gp"], total_gp_budget)
        last_year_gp_share = safe_div(last_year["gp"], total_gp_last_year)
        rows.append(
            [
                key,
                money_m(actual["gmv"]),
                pct(actual_gmv_share),
                pct(budget_gmv_share),
                pct(last_year_gmv_share),
                signed_pp(actual_gmv_share - budget_gmv_share),
                signed_pp(actual_gmv_share - last_year_gmv_share),
                pct(actual_gp_share),
                pct(budget_gp_share),
                pct(last_year_gp_share),
                signed_pp(actual_gp_share - budget_gp_share),
                signed_pp(actual_gp_share - last_year_gp_share),
                pct(rates(actual)["gross_margin"]),
            ]
        )
    return markdown_table(
        [
            "维度",
            "实际GMV",
            "实际GMV占比",
            "预算占比",
            "去年同期占比",
            "较预算",
            "同比",
            "实际GP占比",
            "预算GP占比",
            "去年同期GP占比",
            "较预算",
            "同比",
            "实际毛利率",
        ],
        rows,
    )


def adverse_drivers(actual: dict[str, float], budget: dict[str, float]) -> list[str]:
    comparisons = [
        ("GMV不足", budget["gmv"] - actual["gmv"]),
        ("折扣超支", actual["discount"] - budget["discount"]),
        ("退款超支", actual["refund"] - budget["refund"]),
        ("采购超支", actual["purchase"] - budget["purchase"]),
        ("物流超支", actual["ship"] - budget["ship"]),
        ("广告超支", actual["ads"] - budget["ads"]),
        ("佣金超支", actual["commission"] - budget["commission"]),
        ("FBA/FBM超支", actual["fba"] - budget["fba"]),
        ("仓储超支", actual["storage"] - budget["storage"]),
        ("其他费用超支", actual["other"] - budget["other"]),
    ]
    drivers = [item for item in comparisons if item[1] > 0]
    drivers.sort(key=lambda item: item[1], reverse=True)
    return [f"{name} {money_m(value)}M" for name, value in drivers[:4]]


def rate_gap(actual_metrics: dict[str, float], ref_metrics: dict[str, float], metric_name: str) -> float:
    return rates(actual_metrics)[metric_name] - rates(ref_metrics)[metric_name]


def infer_loss_reasons(
    actual_metrics: dict[str, float],
    ref_metrics: dict[str, float],
    company_metrics: dict[str, float],
    gmv_target_ratio: float,
) -> str:
    if not ref_metrics["gmv"]:
        actual_rates = rates(actual_metrics)
        company_rates = rates(company_metrics)
        adverse = []
        for metric_name, label in [
            ("discount_rate", "折扣"),
            ("refund_rate", "退款"),
            ("ads_rate", "广告"),
            ("purchase_rate", "采购"),
            ("headship_rate", "头程"),
            ("overseas_rate", "海外仓"),
            ("commission_rate", "佣金"),
            ("delivery_rate", "配送费"),
            ("storage_rate", "仓储费"),
            ("other_rate", "其他费用"),
        ]:
            gap = actual_rates[metric_name] - company_rates[metric_name]
            if gap > 0.005:
                adverse.append((gap, f"{label}费率高于公司均值 {signed_pp(gap)}"))
        adverse.sort(key=lambda item: item[0], reverse=True)
        main_judgement = "缺少预算，按实际表现识别为异常亏损"
        lead_parts = [text for _, text in adverse[:3]]
        supplements = [f"毛利率 {pct(actual_rates['gross_margin'])}"]
        if actual_rates["refund_rate"] <= company_rates["refund_rate"] + 0.005:
            supplements.append("退款不是主因")
        return "；".join(
            [main_judgement]
            + ([("主拖累=" + "、".join(lead_parts))] if lead_parts else [])
            + ([("补充=" + "；".join(supplements[:2]))] if supplements else [])
        )

    bridge = gross_profit_bridge(actual_metrics, ref_metrics)
    adverse_rows = sorted(
        [row for row in bridge["rows"] if row["impact"] < -5_000],
        key=lambda row: row["impact"],
    )
    volume_effect = bridge["volume_effect"]
    unit_effect = bridge["unit_economics_effect"]

    if abs(unit_effect) > abs(volume_effect) * 1.3:
        main_judgement = "单位经济恶化主导"
    elif abs(volume_effect) > abs(unit_effect) * 1.3:
        main_judgement = "销量不足主导"
    else:
        main_judgement = "销量与单位经济共同拖累"

    lead_parts = []
    for row in adverse_rows[:3]:
        lead_parts.append(f"{row['factor']} {money_m(row['impact'])}M")

    supplemental = []
    if ref_metrics["gmv"]:
        if gmv_target_ratio < 0.9:
            supplemental.append(f"GMV达成 {ratio(actual_metrics['gmv'], ref_metrics['gmv'])}")
        elif gmv_target_ratio >= 1:
            supplemental.append(f"GMV已达预算 {ratio(actual_metrics['gmv'], ref_metrics['gmv'])}")

    actual_rates = rates(actual_metrics)
    ref_rates = rates(ref_metrics) if ref_metrics["gmv"] else rates(company_metrics)
    if actual_rates["refund_rate"] - ref_rates["refund_rate"] > 0.005:
        supplemental.append(f"退款率偏高 {signed_pp(actual_rates['refund_rate'] - ref_rates['refund_rate'])}")
    else:
        supplemental.append("退款不是主因")

    if actual_metrics["gp"] < 0 and actual_rates["gross_margin"] < 0:
        supplemental.append(f"毛利率 {pct(actual_rates['gross_margin'])}")

    parts = [main_judgement]
    if lead_parts:
        parts.append("主拖累=" + "、".join(lead_parts))
    if supplemental:
        parts.append("补充=" + "；".join(supplemental[:2]))
    return "；".join(parts)


def loss_table(
    actual_models: dict[str, dict[str, float]],
    budget_models: dict[str, dict[str, float]],
    company_metrics: dict[str, float],
    model_info: dict[str, dict[str, str]],
    limit: int,
) -> str:
    rows = []
    filtered = [
        (model, metrics)
        for model, metrics in actual_models.items()
        if model not in IGNORED_MODELS
        and (metrics["gmv"] >= 100_000 or (metrics["gp"] < 0 and metrics["gmv"] >= 50_000))
    ]
    filtered.sort(key=lambda item: item[1]["gp"])
    for model, actual in filtered[:limit]:
        budget = budget_models.get(model, new_metrics())
        info = model_info.get(model, {})
        actual_rates = rates(actual)
        budget_rates = rates(budget)
        gmv_attain = safe_div(actual["gmv"], budget["gmv"]) if budget["gmv"] else 0.0
        reason = infer_loss_reasons(actual, budget, company_metrics, gmv_attain)
        rows.append(
            [
                model,
                info.get("product_line", "未知"),
                info.get("category", "未知"),
                money_m(actual["gmv"]),
                money_m(actual["gp"]),
                pct(actual_rates["gross_margin"]),
                pct(budget_rates["gross_margin"]) if budget["gmv"] else "-",
                ratio(actual["gmv"], budget["gmv"]) if budget["gmv"] else "-",
                pct(actual_rates["refund_rate"]),
                pct(actual_rates["ads_rate"]),
                reason,
            ]
        )
    return markdown_table(
        ["MODEL", "产品线", "三级分类", "GMV", "销售毛利", "实际毛利率", "预算毛利率", "GMV达成", "退款率", "广告费率", "异常归因"],
        rows,
    )


def refund_table(
    actual_groups: dict[str, dict[str, float]],
    last_year_groups: dict[str, dict[str, float]],
    limit: int = 8,
    min_gmv: float = 500_000,
) -> str:
    filtered = [
        (key, metrics)
        for key, metrics in actual_groups.items()
        if metrics["gmv"] >= min_gmv
    ]
    filtered.sort(key=lambda item: item[1]["refund"], reverse=True)
    if not filtered:
        filtered = sorted(actual_groups.items(), key=lambda item: item[1]["refund"], reverse=True)[:limit]
    rows = []
    for key, metrics in filtered[:limit]:
        refund_rate = rates(metrics)["refund_rate"]
        last_year_metrics = last_year_groups.get(key, new_metrics())
        last_year_refund_rate = rates(last_year_metrics)["refund_rate"]
        rows.append(
            [
                key,
                money_m(metrics["gmv"]),
                money_m(metrics["refund"]),
                pct(refund_rate),
                pct(last_year_refund_rate) if last_year_metrics["gmv"] else "-",
                signed_pp(refund_rate - last_year_refund_rate) if last_year_metrics["gmv"] else "-",
                pct(rates(metrics)["gross_margin"]),
            ]
        )
    return markdown_table(
        ["维度", "GMV", "退款额", "本期退款率", "去年同期退款率", "同比变动", "毛利率"],
        rows,
    )


def fee_channel_table(
    actual_groups: dict[str, dict[str, float]],
    budget_groups: dict[str, dict[str, float]],
    nov23_groups: dict[str, dict[str, float]],
    limit: int = 8,
) -> str:
    keys = sorted(actual_groups, key=lambda key: actual_groups[key]["gmv"], reverse=True)
    rows = []
    for key in keys[:limit]:
        actual = actual_groups[key]
        budget = budget_groups.get(key, new_metrics())
        last_year = nov23_groups.get(key, new_metrics())
        actual_rates = rates(actual)
        budget_rates = rates(budget)
        last_year_rates = rates(last_year)
        rows.append(
            [
                key,
                money_m(actual["gmv"]),
                pct(actual_rates["gross_margin"]),
                pp(actual_rates["gross_margin"] - budget_rates["gross_margin"]) if budget["gmv"] else "-",
                pp(actual_rates["gross_margin"] - last_year_rates["gross_margin"]) if last_year["gmv"] else "-",
                pct(actual_rates["ads_rate"]),
                pp(actual_rates["ads_rate"] - budget_rates["ads_rate"]) if budget["gmv"] else "-",
                pp(actual_rates["ads_rate"] - last_year_rates["ads_rate"]) if last_year["gmv"] else "-",
                pct(actual_rates["refund_rate"]),
                pp(actual_rates["refund_rate"] - budget_rates["refund_rate"]) if budget["gmv"] else "-",
            ]
        )
    return markdown_table(
        ["渠道", "GMV", "毛利率", "较预算", "较23年11月", "广告费率", "较预算", "较23年11月", "退款率", "较预算"],
        rows,
    )


def build_report(analyzer: Analyzer) -> str:
    actual_mtd = analyzer.totals[("actual", "mtd")]
    budget_mtd = analyzer.totals[("budget", "mtd")]
    actual_ytd = analyzer.totals[("actual", "ytd")]
    budget_ytd = analyzer.totals[("budget", "ytd")]
    actual_nov23 = analyzer.totals[("actual", "nov23")]
    actual_ytd23 = analyzer.totals[("actual", "ytd23")]

    actual_mtd_rates = rates(actual_mtd)
    budget_mtd_rates = rates(budget_mtd)
    actual_ytd_rates = rates(actual_ytd)
    budget_ytd_rates = rates(budget_ytd)
    actual_nov23_rates = rates(actual_nov23)
    mtd_gmv_pv = gmv_price_volume_analysis(actual_mtd, budget_mtd)
    ytd_gmv_pv = gmv_price_volume_analysis(actual_ytd, budget_ytd)
    mtd_gp_bridge = gross_profit_bridge(actual_mtd, budget_mtd)
    ytd_gp_bridge = gross_profit_bridge(actual_ytd, budget_ytd)
    mtd_adverse, mtd_favorable = summarize_bridge(mtd_gp_bridge)
    ytd_adverse, ytd_favorable = summarize_bridge(ytd_gp_bridge)

    matched_mtd_share, unmatched_mtd_gmv = analyzer.matched_gmv_share("mtd")
    matched_ytd_share, unmatched_ytd_gmv = analyzer.matched_gmv_share("ytd")

    channel_mtd_actual = analyzer.groups[("actual", "mtd", "channel")]
    channel_mtd_budget = analyzer.groups[("budget", "mtd", "channel")]
    channel_nov23_actual = analyzer.groups[("actual", "nov23", "channel")]
    channel_ytd_actual = analyzer.groups[("actual", "ytd", "channel")]
    channel_ytd_budget = analyzer.groups[("budget", "ytd", "channel")]
    channel_ytd23_actual = analyzer.groups[("actual", "ytd23", "channel")]
    region_ytd_actual = analyzer.groups[("actual", "ytd", "region")]
    region_ytd_budget = analyzer.groups[("budget", "ytd", "region")]
    region_ytd23_actual = analyzer.groups[("actual", "ytd23", "region")]
    category_mtd_actual = analyzer.groups[("actual", "mtd", "category")]
    category_mtd_budget = analyzer.groups[("budget", "mtd", "category")]
    category_nov23_actual = analyzer.groups[("actual", "nov23", "category")]
    refund_category_mtd = analyzer.groups[("actual", "mtd", "category")]
    refund_region_mtd = analyzer.groups[("actual", "mtd", "region")]
    refund_category_nov23 = analyzer.groups[("actual", "nov23", "category")]
    refund_region_nov23 = analyzer.groups[("actual", "nov23", "region")]
    channel_region_mtd_actual = aggregate_composite_groups(analyzer, "actual", "mtd", ("channel", "region"))
    channel_region_mtd_budget = aggregate_composite_groups(analyzer, "budget", "mtd", ("channel", "region"))
    category_channel_mtd_actual = aggregate_composite_groups(analyzer, "actual", "mtd", ("category", "channel"))
    category_channel_mtd_budget = aggregate_composite_groups(analyzer, "budget", "mtd", ("category", "channel"))

    low_margin_channel = safe_min_metric(channel_mtd_actual, 1_000_000, "gross_margin")
    high_refund_category = safe_max_metric(refund_category_mtd, 500_000, "refund_rate")
    worst_region_margin = safe_min_metric(analyzer.groups[("actual", "mtd", "region")], 1_000_000, "gross_margin")

    ctx = period_context()
    month_label = str(ctx["month_label"])
    current_month_text = str(ctx["current_month_text"])
    last_year_month_text = str(ctx["last_year_month_text"])
    current_ytd_text = str(ctx["current_ytd_text"])

    lines = []
    lines.append(f"# {current_month_text}产品销售经营分析报告")
    lines.append("")
    lines.append(
        f"口径说明：基于 `预算数据源` 与 `实际数据源`，分析范围为公司整体产品销售表现；当月指 {current_month_text}，YTD 指 {current_ytd_text}，金额单位为 USD million。物流拆分口径中，预算侧“头程”按 `预算头程成本+预算二程配送费+预算关税` 汇总；实际侧无单独 FBM 字段，配送费暂按 `订单FBA费用` 口径呈现。"
    )
    lines.append("")
    lines.append("## 一、经营结论")
    lines.append("")
    lines.append(
        f"1. {month_label} GMV 为 {money_m(actual_mtd['gmv'])}M，较预算少 {money_m(budget_mtd['gmv'] - actual_mtd['gmv'])}M，达成率 {ratio(actual_mtd['gmv'], budget_mtd['gmv'])}；销售毛利 {money_m(actual_mtd['gp'])}M，仅为预算的 {ratio(actual_mtd['gp'], budget_mtd['gp'])}，毛利率 {pct(actual_mtd_rates['gross_margin'])}，较预算变动 {signed_pp(actual_mtd_rates['gross_margin'] - budget_mtd_rates['gross_margin'])}。"
    )
    lines.append(
        f"2. YTD GMV 为 {money_m(actual_ytd['gmv'])}M，达成全年同期预算 {ratio(actual_ytd['gmv'], budget_ytd['gmv'])}；YTD 销售毛利 {money_m(actual_ytd['gp'])}M，达成率 {ratio(actual_ytd['gp'], budget_ytd['gp'])}，毛利率 {pct(actual_ytd_rates['gross_margin'])}，较预算变动 {signed_pp(actual_ytd_rates['gross_margin'] - budget_ytd_rates['gross_margin'])}。"
    )
    lines.append(
        f"3. {month_label}同比 {last_year_month_text}，GMV 同比 {signed_pct(yoy_change(actual_mtd['gmv'], actual_nov23['gmv']))}，销售毛利同比 {signed_pct(yoy_change(actual_mtd['gp'], actual_nov23['gp']))}；毛利率同比变动 {signed_pp(actual_mtd_rates['gross_margin'] - actual_nov23_rates['gross_margin'])}。收入与利润趋势并不同步，需要继续结合价差/量差和费用结构拆解判断。"
    )
    lines.append(
        f"4. 预算匹配度方面，{month_label}按 `渠道+地区+MODEL` 口径可匹配预算的实际 GMV 占比 {pct(matched_mtd_share)}，仍有 {money_m(unmatched_mtd_gmv)}M GMV 落在未完整匹配预算的组合中；YTD 匹配占比 {pct(matched_ytd_share)}。预算覆盖不足会放大结构差异判断误差，但不改变整体利润承压结论。"
    )
    lines.append(
        f"5. 从利润桥看，{month_label}毛利较预算少 {money_m(abs(mtd_gp_bridge['total_effect']))}M，其中量差影响 {money_m(mtd_gp_bridge['volume_effect'])}M，单位经济差影响 {money_m(mtd_gp_bridge['unit_economics_effect'])}M。单位经济层面的主要利空因素是："
        + "，".join(mtd_adverse[:4])
        + "。"
    )
    lines.append("")
    lines.append("## 二、损益分析")
    lines.append("")
    lines.append(f"### 1. {month_label}：实际 vs 预算 vs 去年同期")
    lines.append("")
    lines.append(combined_metrics_table(actual_mtd, budget_mtd, actual_nov23))
    lines.append("")
    lines.append("### 2. YTD：实际 vs 预算 vs 去年同期")
    lines.append("")
    lines.append(combined_metrics_table(actual_ytd, budget_ytd, actual_ytd23))
    lines.append("")
    lines.append("## 三、利润因素分析（价差/量差）")
    lines.append("")
    lines.append(f"### 1. {month_label} GMV 价差与量差")
    lines.append("")
    lines.append(gmv_pv_table(actual_mtd, budget_mtd))
    lines.append("")
    lines.append(
        f"{month_label} GMV 差异 {money_m(mtd_gmv_pv['total_effect'])}M 中，量差影响 {money_m(mtd_gmv_pv['volume_effect'])}M，价差影响 {money_m(mtd_gmv_pv['price_effect'])}M。说明当月收入缺口同时来自销量和单价/结构，其中量差影响更大。"
    )
    lines.append("")
    lines.append(f"### 2. {month_label}毛利利润桥")
    lines.append("")
    lines.append(gross_profit_bridge_table(actual_mtd, budget_mtd))
    lines.append("")
    lines.append(
        f"{month_label}利润差异中，量差拖累 {money_m(mtd_gp_bridge['volume_effect'])}M；单位经济差再拖累 {money_m(mtd_gp_bridge['unit_economics_effect'])}M。主要利空因素为 {('，'.join(mtd_adverse[:4]) if mtd_adverse else '无明显利空')}；主要利好因素为 {('，'.join(mtd_favorable[:3]) if mtd_favorable else '无明显利好')}。"
    )
    lines.append(
        "其中佣金单价上升需要谨慎解读。佣金本质上由平台抽佣比例、平台结构变化以及退货后的佣金损失共同决定，更适合视为结构结果项，而不是单独的费控失效。"
    )
    lines.append("")
    lines.append("### 3. YTD 毛利利润桥")
    lines.append("")
    lines.append(gross_profit_bridge_table(actual_ytd, budget_ytd))
    lines.append("")
    lines.append(
        f"YTD 毛利较预算少 {money_m(abs(ytd_gp_bridge['total_effect']))}M，其中量差影响 {money_m(ytd_gp_bridge['volume_effect'])}M，单位经济差影响 {money_m(ytd_gp_bridge['unit_economics_effect'])}M。YTD 主要利空因素为 {('，'.join(ytd_adverse[:4]) if ytd_adverse else '无明显利空')}；主要利好因素为 {('，'.join(ytd_favorable[:3]) if ytd_favorable else '无明显利好')}。"
    )
    lines.append("")
    lines.append(f"### 4. {month_label}渠道毛利差异拆解")
    lines.append("")
    lines.append(channel_profit_variance_table(channel_mtd_actual, channel_mtd_budget, limit=8))
    lines.append("")
    lines.append(f"### 5. {month_label}渠道×区域毛利差异拆解")
    lines.append("")
    lines.append(
        composite_profit_variance_table(
            channel_region_mtd_actual,
            channel_region_mtd_budget,
            f"{month_label}渠道×区域毛利差异拆解",
            limit=12,
        ).split("\n", 2)[2]
    )
    lines.append("")
    lines.append(f"### 6. {month_label}类目×平台(渠道)毛利差异拆解")
    lines.append("")
    lines.append(
        composite_profit_variance_table(
            category_channel_mtd_actual,
            category_channel_mtd_budget,
            f"{month_label}类目×平台(渠道)毛利差异拆解",
            limit=12,
        ).split("\n", 2)[2]
    )
    lines.append("")
    lines.append("## 四、产品结构分析")
    lines.append("")
    lines.append("### 1. 渠道结构（YTD）")
    lines.append("")
    lines.append(
        top_groups_table(
            channel_ytd_actual,
            channel_ytd_budget,
            channel_ytd23_actual,
            actual_ytd,
            budget_ytd,
            actual_ytd23,
            limit=8,
        )
    )
    lines.append("")
    lines.append("### 2. 区域结构（YTD）")
    lines.append("")
    lines.append(
        top_groups_table(
            region_ytd_actual,
            region_ytd_budget,
            region_ytd23_actual,
            actual_ytd,
            budget_ytd,
            actual_ytd23,
            limit=8,
        )
    )
    lines.append("")
    lines.append(f"### 3. 三级分类结构（{month_label}）")
    lines.append("")
    lines.append(
        top_groups_table(
            category_mtd_actual,
            category_mtd_budget,
            category_nov23_actual,
            actual_mtd,
            budget_mtd,
            actual_nov23,
            limit=10,
        )
    )
    lines.append("")
    structure_notes: list[str] = []
    if low_margin_channel:
        structure_notes.append(
            f"低毛利渠道中，{low_margin_channel[0]} 毛利率仅 {pct(rates(low_margin_channel[1])['gross_margin'])}"
        )
    if high_refund_category:
        structure_notes.append(
            f"退款压力最大的三级分类为 {high_refund_category[0]}，退款率 {pct(rates(high_refund_category[1])['refund_rate'])}"
        )
    if worst_region_margin:
        structure_notes.append(
            f"区域中毛利率最低的是 {worst_region_margin[0]}，仅 {pct(rates(worst_region_margin[1])['gross_margin'])}"
        )
    if structure_notes:
        lines.append(f"结构观察：{month_label}" + "；".join(structure_notes) + "。")
    else:
        lines.append(f"结构观察：{month_label}未筛出满足样本阈值的低毛利渠道/区域或高退款类目，需结合更低阈值或明细样本复核。")
    lines.append("")
    lines.append("## 五、产品实际费率分析")
    lines.append("")
    lines.append(f"### 1. {month_label}渠道费率表现")
    lines.append("")
    lines.append(fee_channel_table(channel_mtd_actual, channel_mtd_budget, channel_nov23_actual, limit=8))
    lines.append("")
    lines.append(
        f"{month_label}公司整体广告费率 {pct(actual_mtd_rates['ads_rate'])}，较预算变动 {signed_pp(actual_mtd_rates['ads_rate'] - budget_mtd_rates['ads_rate'])}，较 {last_year_month_text}变动 {signed_pp(actual_mtd_rates['ads_rate'] - actual_nov23_rates['ads_rate'])}；平台费率 {pct(actual_mtd_rates['platform_fee_rate'])}，较预算变动 {signed_pp(actual_mtd_rates['platform_fee_rate'] - budget_mtd_rates['platform_fee_rate'])}；物流费率 {pct(actual_mtd_rates['logistics_rate'])}，较预算变动 {signed_pp(actual_mtd_rates['logistics_rate'] - budget_mtd_rates['logistics_rate'])}。费用率变化直接影响毛利率表现。"
    )
    lines.append("")
    lines.append("## 六、产品实际退款率分析")
    lines.append("")
    lines.append(
        f"{month_label}公司整体退款率为 {pct(actual_mtd_rates['refund_rate'])}，较预算变动 {signed_pp(actual_mtd_rates['refund_rate'] - budget_mtd_rates['refund_rate'])}，较 {last_year_month_text}变动 {signed_pp(actual_mtd_rates['refund_rate'] - actual_nov23_rates['refund_rate'])}。退款率分析应优先看各品类和区域相对去年同期的变化，而不是简单对比公司均值。"
    )
    lines.append("")
    lines.extend(channel_refund_sections(analyzer))
    lines.append("## 七、亏损异常产品")
    lines.append("")
    lines.append(f"### 1. {month_label}亏损/低毛利重点 MODEL")
    lines.append("")
    lines.append(
        loss_table(
            analyzer.models[("actual", "mtd")],
            analyzer.models[("budget", "mtd")],
            actual_mtd,
            analyzer.model_info,
            limit=10,
        )
    )
    lines.append("")
    lines.append("### 2. YTD 亏损/低毛利重点 MODEL")
    lines.append("")
    lines.append(
        loss_table(
            analyzer.models[("actual", "ytd")],
            analyzer.models[("budget", "ytd")],
            actual_ytd,
            analyzer.model_info,
            limit=10,
        )
    )
    lines.append("")
    lines.append("## 八、主要经营问题与建议")
    lines.append("")
    lines.append(
        f"1. 利润短板大于规模短板。{month_label}GMV 只比预算少 {money_m(budget_mtd['gmv'] - actual_mtd['gmv'])}M，但销售毛利比预算少 {money_m(budget_mtd['gp'] - actual_mtd['gp'])}M，说明经营压力核心在费率和成本率，而不是单纯销量不足。"
    )
    lines.append(
        f"2. 利润因素拆解显示，{month_label}销量不足拖累毛利 {money_m(mtd_gp_bridge['volume_effect'])}M，单位经济变化影响 {money_m(mtd_gp_bridge['unit_economics_effect'])}M。当前更需要优先治理的是售价/结构、广告投放和物流单价等单位经济问题。"
    )
    lines.append(
        "3. 佣金因素不应孤立解读。佣金单价上升更多反映平台结构变化和退货损失，不建议将其直接视为独立经营失误；真正需要优先治理的仍是售价/结构、广告投放效率与物流成本。"
    )
    lines.append(
        f"4. 费用结构变化主要来自采购和投放。{month_label}广告费率同比{pp_change_text(actual_mtd_rates['ads_rate'] - actual_nov23_rates['ads_rate'])}，平台费率同比{pp_change_text(actual_mtd_rates['platform_fee_rate'] - actual_nov23_rates['platform_fee_rate'])}，采购费率同比{pp_change_text(actual_mtd_rates['purchase_rate'] - actual_nov23_rates['purchase_rate'])}；虽然物流费率同比{pp_change_text(actual_mtd_rates['logistics_rate'] - actual_nov23_rates['logistics_rate'])}，但不足以完全对冲毛利率同比变化 {abs(actual_mtd_rates['gross_margin'] - actual_nov23_rates['gross_margin']) * 100:.2f}pp。"
    )
    focus_parts = [item[0] for item in (low_margin_channel, worst_region_margin, high_refund_category) if item]
    if focus_parts:
        lines.append(
            f"5. 结构性问题仍然明显。部分低毛利渠道、低毛利区域和高退款类目占据了较高 GMV 权重，形成“收入放大、利润不放大”的结构。尤其是 {'、'.join(focus_parts)}，需要作为后续经营修复重点。"
        )
    else:
        lines.append(
            "5. 结构性问题仍然明显，但当前按默认阈值未筛出稳定的低毛利渠道/区域或高退款类目，需要结合更低颗粒度样本继续复核。"
        )
    lines.append(
        "6. 预算管理需要补齐产品组合口径。存在一定未完整匹配预算的 `渠道+地区+MODEL` 组合，建议后续将预算与实际统一到相同产品维度并补齐新增 SKU/渠道预算，以便更准确地定位结构偏差。"
    )
    lines.append(
        "7. 建议围绕亏损 MODEL 做逐项复盘：先核查定价与折扣策略，再检查广告投放效率和退款成因，最后按区域/渠道判断是否需要做清库存、停售或转仓。"
    )
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    global TARGET_YEAR, TARGET_MONTH
    project_root = Path.cwd()
    parser = argparse.ArgumentParser(description="Generate operating analysis report from demo.xlsx")
    parser.add_argument(
        "--input",
        default=str(project_root / "data" / "demo.xlsx"),
        help="Path to the input workbook",
    )
    parser.add_argument(
        "--run-root",
        default=None,
        help="Root directory for this analysis run. If omitted, a new timestamped run directory is created under output/runs/.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Path to the generated markdown report. Defaults to <run_root>/final/operating-analysis.md",
    )
    args = parser.parse_args()

    run_root = resolve_run_root(project_root, args.run_root)
    TARGET_YEAR, TARGET_MONTH = detect_target_period(Path(args.input))
    analyzer = Analyzer(Path(args.input))
    analyzer.run()
    report = build_report(analyzer)
    output_path = Path(args.output) if args.output else stage_dir(run_root, "final") / "operating-analysis.md"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(report, encoding="utf-8")
    print(run_root)


if __name__ == "__main__":
    main()
